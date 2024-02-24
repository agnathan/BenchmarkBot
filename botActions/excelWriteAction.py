import os
import json
import glob

import openpyxl
from openpyxl.styles import Alignment

# Import project classes
from botActions.action import Action


class ExcelWriteAction(Action):
    def run(self, config, result):
        print("\t\033[92mExcelWriteAction running ...\033[0m")
        print("\033[93m\t Activated by: ", result, "\033[0m")

        def read_nlp_json_files(directory):
            # Pattern to match files ending with '-nlp.json'
            pattern = os.path.join(directory, "*-nlp.json")

            # Find all files in the directory matching the pattern
            file_paths = glob.glob(pattern)

            # Initialize an array to hold the contents of all files
            contents = []

            # Iterate over the list of file paths
            for file_path in file_paths:
                # Open and read the JSON file
                with open(file_path, "r") as file:
                    data = json.load(file)
                    # Append the data to the contents array
                    contents.append(data)

            return contents

        def buildEl(el, indexCol, index, xlength, benchmark, dp):
            # print(el)
            # print(indexCol, index, indexCol==index)
            if indexCol == index and benchmark == dp["Benchmark"]:
                return dp["Score"]
            else:
                return el

        # Returns the original row elements except
        def buildRow(currentRow, productYIndex, arrayWidth, dp):
            # print("currentRow: ", currentRow)
            # print("productYIndex: ", productYIndex)
            # print("arrayWidth: ", arrayWidth)

            newRow = []
            if len(currentRow) == 0 or currentRow[0] == dp["Benchmark"]:
                for i in range(arrayWidth):
                    if i == 0:
                        newRow.append(dp["Benchmark"])
                    elif i == 1:
                        newRow.append(dp["Type"])
                    elif i == 2:
                        newRow.append(dp["Metric"])
                    elif i == productYIndex:
                        newRow.append(dp["Score"])
                    elif i < len(currentRow):
                        newRow.append(currentRow[i])
                    else:
                        newRow.append("")
            else:
                for i in range(arrayWidth):
                    if i < len(currentRow):
                        newRow.append(currentRow[i])
                    else:
                        newRow.append("")
            return newRow

        def getCol(dp):
            col = ""
            if "Product" in dp:
                product = dp["Product"]
                col = product
            if "Processor" in dp:
                processor = dp["Processor"]
                col = processor
            if "Product" in dp and "Processor" in dp:
                col = product + "\n" + processor
            return col

        def dataPointHeaders(dp):
            return ["Benchmark", "Type", "Metric", getCol(dp)]

        def mergeHeader(h1, h2):
            union_array = []
            [union_array.append(item) for item in h1 + h2 if item not in union_array]
            return union_array

        def getHeadersAndData(inputArray):
            try:
                headers = inputArray[0]
            except:
                headers = ["Benchmark", "Type", "Metric"]

            data = inputArray[1:]
            return (headers, data)

        def getYColumnIndex(newHeaders, dp):
            # Find the index of a Product/Processor in the array
            # or return the length of the array
            xlength = len(newHeaders)

            # if the Col is already in the results then return the x index (the column number)
            try:
                productYIndex = newHeaders.index(getCol(dp))
            except ValueError:
                # This should never execute ... better error handling needed
                productYIndex = len(newHeaders)
            return productYIndex

        def getXBenchmarkIndex(oldData, dp):
            try:
                benchmarkArray = [item[0] for item in oldData]
                xIndex = benchmarkArray.index(dp["Benchmark"])
            except:
                xIndex = None
            return xIndex

        def addDataPoint(inputArray, dp):
            # print("\nEntering addDataPoint...")
            # print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            # print("inputArray: ", inputArray)
            # print("dp: ", dp)
            # print("===============================================================")
            oldHeaders, oldData = getHeadersAndData(inputArray)
            # Declare the new data array that will be returned
            newData = []

            # Find the headers from the new data point and merge them
            # with the current headers
            dpHeaders = dataPointHeaders(dp)
            newHeaders = mergeHeader(oldHeaders, dpHeaders)

            # Search the new headers array for the product/processor column and return its index
            productYIndex = getYColumnIndex(newHeaders, dp)
            # print("productYIndex: ", productYIndex)

            benchmarkXIndex = getXBenchmarkIndex(oldData, dp)
            # print("benchmarkXIndex: ", benchmarkXIndex)

            arrayWidth = len(newHeaders)
            # Search through the array for a row that contains the current benchmark
            # or return the benchmarkXIndex as the height of the array
            newRow = []

            # Loop through each row in the data and if the be
            for index, row in enumerate(oldData):
                # Is the current row the same benchmark as the incoming data point
                newRow = buildRow(row, productYIndex, arrayWidth, dp)
                newData.append(newRow)

            # benchmarkXIndex is zero when the dp benchmark is not found in the inputArray benchmarks
            if benchmarkXIndex == None:
                newRow = buildRow([], productYIndex, arrayWidth, dp)
                newData.append(newRow)

            newResults = [newHeaders] + newData
            return newResults

        # print(config)
        cached_dirname_path = os.path.dirname(result)
        cached_results_path = os.path.join(
            cached_dirname_path, "all-excel-results.json"
        )
        # print("\033[93m Activated by: ", cached_dirname_path, "\033[0m")
        # if os.path.exists(cached_file_path):
        #     print(
        #         f"NamedEntityRecognitionAction results already cached at {cached_file_path}"
        #     )
        #     return cached_file_path
        data = read_nlp_json_files(cached_dirname_path)

        points = []
        # Get all of the data points
        for nlpFile in data:
            for dp in nlpFile:
                points = addDataPoint((points), dp)

        # Define the Upper-Left Excel Cell to write
        start_cell = "A22"
        row = 22
        outputfile = config["outputfile"]

        # Load the workbook if it exists, otherwise create a new one
        try:
            workbook = openpyxl.load_workbook(outputfile)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        sheet = workbook[config["sheet"]]

        # Clear data in the current worksheet
        # Calculate the max row and column to know the range to clear
        max_row = sheet.max_row
        max_column = sheet.max_column

        # Iterate over each row and column to clear the cells below the specified row
        for i in range(row + 1, max_row + 1):
            for j in range(1, max_column + 1):
                sheet.cell(row=i, column=j).value = None

        # Determine the starting row and column from the start_cell
        start_column = openpyxl.utils.column_index_from_string(start_cell[0])
        start_row = int(start_cell[1:])

        # print("points: ", points)
        # Write the two-dimensional array data to the sheet starting at the given cell
        for i, row in enumerate(points, start=start_row):
            for j, value in enumerate(row, start=start_column):
                cell = sheet.cell(row=i, column=j)
                cell.value = value
                cell.alignment = Alignment(wrapText=True)
        try:
            # Save the workbook
            workbook.save(outputfile)
        except PermissionError:
            print(
                "\033[1;31mPlease, close the inteview-question.xlsx so the benchmarkbot can write to it\033[0m"
            )
